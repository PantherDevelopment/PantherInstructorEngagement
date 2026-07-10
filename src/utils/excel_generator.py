"""
Excel Report Generator
Two distinct column structures for current_week vs entire_term reports
"""

from datetime import datetime
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:

    MAROON     = "770000"
    LIGHT_GRAY = "F2F2F2"
    WHITE      = "FFFFFF"

    # Each entry: (header, column_width, data_key)
    WEEK_COLUMNS = [
        ('Course Code',                               18, 'course_code'),
        ('Course Name',                               32, 'course_name'),
        ('Instructor',                                26, 'instructor_name'),
        ('Total Activity Time (Term)',                24, 'total_activity_time'),
        ('Last Login Date (ET)',                      24, 'last_login_date'),
        ('Discussion Posts: Due >7 Days Ago, Not Graded',   32, 'discussion_not_graded_7days'),
        ('Other Assignments: Due >7 Days Ago, Not Graded', 32, 'other_not_graded_7days'),
        ('Discussion Replies (This Week)',            24, 'discussion_replies_week'),
        ('Discussion Replies (Term Total)',           24, 'discussion_replies_total'),
        ('Announcements (This Week)',                  22, 'announcements_count'),
        ('Instructor Bio Complete',                    22, 'instructor_bio_complete'),
        ('Data Collection Notes',                      35, 'collection_notes'),
    ]

    TERM_COLUMNS = [
        ('Course Code',                               18, 'course_code'),
        ('Course Name',                               32, 'course_name'),
        ('Instructor',                                26, 'instructor_name'),
        ('Total Activity Time (Term)',                24, 'total_activity_time'),
        ('Last Login Date (ET)',                      24, 'last_login_date'),
        ('Discussion Posts: Due >7 Days Ago, Not Graded',   32, 'discussion_not_graded_7days'),
        ('Other Assignments: Due >7 Days Ago, Not Graded', 32, 'other_not_graded_7days'),
        ('Discussion Replies (Term Total)',           24, 'discussion_replies_total'),
        ('Announcements (Term Total)',                 22, 'announcements_count'),
        ('Instructor Bio Complete',                    22, 'instructor_bio_complete'),
        ('Data Collection Notes',                      35, 'collection_notes'),
    ]

    # Keys that should be center-aligned
    CENTER_KEYS = {
        'total_activity_time', 'login_count_week',
        'discussion_replies_week', 'discussion_replies_total',
        'discussion_not_graded_7days', 'other_not_graded_7days',
        'announcements_count', 'instructor_bio_complete'
    }

    def generate_report(self, data: List[Dict], output_path: str, period: str = 'current_week') -> str:
        columns = self.WEEK_COLUMNS if period == 'current_week' else self.TERM_COLUMNS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Instructor Engagement"

        # Column widths
        for col_idx, (_, width, _key) in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Styles
        thin         = Side(style='thin')
        border       = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill  = PatternFill(start_color=self.MAROON, end_color=self.MAROON, fill_type="solid")
        header_font  = Font(bold=True, color=self.WHITE, size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center       = Alignment(horizontal='center')

        # Header row
        ws.row_dimensions[1].height = 50
        for col_idx, (header, _, _key) in enumerate(columns, 1):
            cell           = ws.cell(row=1, column=col_idx)
            cell.value     = header
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = border

        # Data rows
        for row_num, record in enumerate(data, 2):
            fill_color = self.LIGHT_GRAY if row_num % 2 == 0 else self.WHITE
            row_fill   = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            for col_idx, (_header, _width, key) in enumerate(columns, 1):
                cell        = ws.cell(row=row_num, column=col_idx)
                value       = record.get(key, '')
                cell.value  = value
                cell.fill   = row_fill
                cell.border = border
                if key in self.CENTER_KEYS:
                    cell.alignment = center
                # Style N/A cells in italic gray
                if value == 'N/A':
                    cell.font      = Font(italic=True, color="999999", size=10)
                    cell.alignment = center
                # Highlight notes cell red if there were errors
                elif key == 'collection_notes' and record.get(key, 'OK') != 'OK':
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    cell.font = Font(color="CC0000", size=10)

        ws.freeze_panes = 'A2'
        wb.save(output_path)
        return output_path